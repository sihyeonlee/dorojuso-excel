# -*- coding: utf-8 -*-
from urllib.parse import quote_plus, urlencode
from urllib.request import urlopen, Request
from openpyxl.styles import PatternFill, Color
from openpyxl import load_workbook
import json


def get_juso(name):
    url = 'http://www.juso.go.kr/addrlink/addrLinkApi.do'
    queryParams = '?' + urlencode(
        {quote_plus('currentPage'): '1', quote_plus('countPerPage'): '10', quote_plus('resultType'): 'json',
         quote_plus('keyword'): name, quote_plus('confmKey'): '인증키'})

    request = Request(url + queryParams)
    request.get_method = lambda: 'GET'
    response_body = urlopen(request).read()

    root_json = json.loads(response_body)

    if root_json['results']['common']['totalCount'] == '0':
        return False

    for child in root_json['results']['juso']:
        if child['siNm'] != '서울특별시':
            return False
        else:
            dorojuso = child['roadAddr']
            jibunjuso = child['jibunAddr']

            return dorojuso, jibunjuso

filename = "dorojuso.xlsx"

workbook = load_workbook(filename)
sheet = workbook['Sheet1']
index = 2

while True:
    name_cell = "A" + str(index)
    dorojuso_cell = "B" + str(index)
    jibunjuso_cell = "C" + str(index)

    if sheet[name_cell].value == None:
        break

    result = get_juso(sheet[name_cell].value)

    if not result:
        sheet[dorojuso_cell] = "서울시내에 검색된 결과가 없습니다."
        sheet[dorojuso_cell].fill = PatternFill(patternType='solid', fgColor=Color('FF0000'))

    else:
        dorojuso = result[0]
        jibunjuso = result[1]

        sheet[dorojuso_cell] = dorojuso
        sheet[jibunjuso_cell] = jibunjuso

    index = index + 1

workbook.save('result_dorojuso.xlsx')
